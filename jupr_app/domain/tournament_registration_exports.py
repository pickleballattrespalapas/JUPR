from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value) + 2)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 48)


def _write_table(ws, headers: list[str], rows: list[list[Any]], *, start_row: int = 1) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True)
    row_idx = start_row + 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
        row_idx += 1
    _autosize(ws)
    return row_idx


def build_registration_workbook(
    *,
    tournament: dict[str, Any],
    state: dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    summary = state.get("summary", {})
    summary_rows = [
        ["Tournament", tournament.get("name")],
        ["Registrations", summary.get("total_registrations")],
        ["Selections", summary.get("total_selections")],
        ["Confirmed entries", summary.get("confirmed_entries")],
        ["Review entries", summary.get("review_entries")],
        ["Waitlist entries", summary.get("waitlist_entries")],
        ["Needs partner", summary.get("needs_partner_entries")],
        ["Partner missing", summary.get("partner_missing_entries")],
        ["Issues", summary.get("issue_count")],
        ["Blockers", summary.get("blocker_count")],
    ]
    _write_table(ws, ["Metric", "Value"], summary_rows)

    # Registrations sheet
    regs_ws = wb.create_sheet("Registrations")
    reg_headers = [
        "Submitted At",
        "Name",
        "Email",
        "Phone",
        "DUPR",
        "Doubles Skill",
        "Singles Skill",
        "Age",
        "Age Bracket",
        "Gender",
        "Payment Status",
        "Notes",
    ]
    reg_rows = []
    for reg in state.get("registrations", []):
        reg_rows.append(
            [
                reg.get("submitted_at"),
                reg.get("display_name"),
                reg.get("email"),
                reg.get("phone"),
                reg.get("dupr_id"),
                reg.get("doubles_skill"),
                reg.get("singles_skill"),
                reg.get("age"),
                reg.get("age_bracket"),
                reg.get("gender"),
                reg.get("payment_status"),
                reg.get("notes"),
            ]
        )
    _write_table(regs_ws, reg_headers, reg_rows)

    # Issues sheet
    issues_ws = wb.create_sheet("Issues")
    issue_headers = ["Severity", "Type", "Event Option ID", "Registration ID", "Selection ID", "Message"]
    issue_rows = []
    for issue in state.get("issues", []):
        issue_rows.append(
            [
                issue.get("severity"),
                issue.get("issue_type"),
                issue.get("event_option_id"),
                issue.get("registration_id"),
                issue.get("selection_id"),
                issue.get("message"),
            ]
        )
    _write_table(issues_ws, issue_headers, issue_rows)

    # Partner board sheet
    partner_ws = wb.create_sheet("Partner Board")
    partner_headers = ["Day", "Event", "Player", "Email", "Skill", "Age", "Note"]
    partner_rows = []
    for row in state.get("partner_board", []):
        player = row.get("player") or {}
        partner_rows.append(
            [
                row.get("event_day_label"),
                row.get("event_label"),
                player.get("display_name"),
                player.get("email") if row.get("show_contact_email") else None,
                player.get("skill"),
                player.get("age"),
                row.get("note"),
            ]
        )
    _write_table(partner_ws, partner_headers, partner_rows)

    # Event sheets
    for roster in state.get("event_rosters", []):
        event_label = str(roster.get("event_label") or "Roster")
        safe_title = event_label[:28] if len(event_label) > 28 else event_label
        safe_title = safe_title or "Roster"
        candidate = safe_title
        suffix = 2
        while candidate in wb.sheetnames:
            candidate = f"{safe_title[:25]}-{suffix}"
            suffix += 1
        roster_ws = wb.create_sheet(candidate)
        roster_headers = ["Status", "Member 1", "Member 1 Email", "Member 2", "Member 2 Email", "Submitted At"]
        roster_rows = []
        for entry in roster.get("entries", []):
            members = entry.get("members") or []
            m1 = members[0] if len(members) > 0 else {}
            m2 = members[1] if len(members) > 1 else {}
            roster_rows.append(
                [
                    entry.get("status"),
                    m1.get("display_name"),
                    m1.get("email"),
                    m2.get("display_name"),
                    m2.get("email"),
                    entry.get("submitted_at"),
                ]
            )
        _write_table(roster_ws, roster_headers, roster_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
